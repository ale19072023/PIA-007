import requests
import json
import mainESP
import openpyxl
from openpyxl.chart import BarChart, Reference
import os
import datetime

def fun_planetas():
    url = "http://swapi.dev/api/planets/"
    response = requests.get(url, verify=False)

    if response.status_code == 200:
        menu_planeta()
    else:
        print("Error al consultar el API")

    print(response.status_code)

def menu_planeta():
    opcion = 0
    print("Informacion de Planetas: ")
    while True:
        print("\n *** Bienvenido al menú de planetas: ***")
        print("1. Buscar planeta por ID")
        print("2. Buscar planeta por nombre")
        print("3. Obtener información de población de planetas")
        print("4. Calcular media poblacional planetaria")
        print("5. Volver al menú principal")
        opcion = int(input("Selecciona una Opcion: "))

        if opcion == 1:
            id_planeta = int(input("Ingresa el ID del planeta: "))
            obtener_planeta_por_id(id_planeta)
        elif opcion == 2:
            nombre_planeta = str(input("Ingresa el nombre del planeta: "))
            obtener_planetas_por_nombre(nombre_planeta)
        elif opcion == 3:
            num_planetas = int(input("Ingresa el número de planetas para obtener información de población: "))
            nombres_planetas, poblaciones_planetas = obtener_datos_poblacion(num_planetas)
            graficar_datos_poblacion(nombres_planetas, poblaciones_planetas)
        elif opcion == 4:
            calcular_media_poblacional()
        elif opcion == 5:
            mainESP.main_menu()
        else:
            print("Opcion no valida, intenta de nuevo")

def obtener_planeta_por_id(id):
    url_planeta = f"https://swapi.dev/api/planets/{id}"
    try:
        response = requests.get(url_planeta)
        response.raise_for_status()
        datos_planeta = json.loads(response.text)
        print(f"Nombre del planeta encontrado: {datos_planeta['name']}\n")
        return datos_planeta
    except requests.exceptions.RequestException as e:
        print(f"Error al obtener datos del planeta {id}: {e}")
        return None

def obtener_planetas_por_nombre(nombre):
    response = requests.get(f"https://swapi.dev/api/planets/?search={nombre}")
    if response.status_code == 200:
        resultado = response.json()
        if resultado["count"] == 1:
            planeta = resultado["results"][0]
            print(f"Nombre: {planeta['name']}")
            print(f"Clima: {planeta['climate']}")
            print(f"Terreno: {planeta['terrain']}")
            print(f"Periodo de rotación: {planeta['rotation_period']} horas")
            print(f"Periodo orbital: {planeta['orbital_period']} días")
            print(f"Población: {planeta['population']} habitantes")
            return planeta
        elif resultado["count"] == 0:
            print(f"No se encontro ningun planeta con el nombre '{nombre}'.")
        else:
            print(f"Se encontraron varios planetas con el nombre '{nombre}'. Por favor especifique su buqueda.")
    else:
        print(f"Hubo un error al realizar la peticion: {response.status_code}")

def obtener_datos_poblacion(num_planetas):
    nombres_planetas = []
    poblaciones_planetas = []

    for i in range(1, num_planetas + 1):
        url_planeta = f"https://swapi.dev/api/planets/{i}"
        response = requests.get(url_planeta)

        if response.status_code == 200:
            datos_planeta = response.json()
            if datos_planeta["population"] == "unknown":
                poblacion_planeta = 0
            else:
                poblacion_planeta = int(datos_planeta["population"].replace(",", ""))

            nombres_planetas.append(datos_planeta["name"])
            poblaciones_planetas.append(poblacion_planeta)
        else:
            print(f"Error al obtener datos del planeta {i}. Código de respuesta HTTP: {response.status_code}")

    return nombres_planetas, poblaciones_planetas

def graficar_datos_poblacion(nombres_planetas, poblaciones_planetas):
    fecha_hora_actual = datetime.datetime.now().strftime("%Y%m%d%H%M%S")
    carpeta_graficas = "graficas"
    if not os.path.exists(carpeta_graficas):
        os.makedirs(carpeta_graficas)

    libro_trabajo = openpyxl.Workbook()
    hoja = libro_trabajo.active
    for i in range(len(nombres_planetas)):
        hoja.cell(row=i+1, column=1, value=nombres_planetas[i])
        hoja.cell(row=i+1, column=2, value=poblaciones_planetas[i])
    grafico = BarChart()
    grafico.type = "col"
    grafico.title = "Población de los planetas de Star Wars"
    grafico.y_axis.title = "Población"
    grafico.x_axis.title = "Planetas"
    datos = Reference(hoja, min_col=2, min_row=1, max_row=len(nombres_planetas))
    etiquetas = Reference(hoja, min_col=1, min_row=1, max_row=len(nombres_planetas))
    grafico.add_data(datos, titles_from_data=True)
    grafico.set_categories(etiquetas)
    grafico.shape = 4
    hoja.add_chart(grafico, "D1")
    libro_trabajo.save(f"{carpeta_graficas}/datos_poblacion_{fecha_hora_actual}.xlsx")

def calcular_media_poblacional():
    num_planetas = int(input("Ingrese el número de planetas para calcular la media poblacional: "))
    _, poblaciones_planetas = obtener_datos_poblacion(num_planetas)
    media_poblacional = sum(poblaciones_planetas) / len(poblaciones_planetas)
    print(f"La media poblacional planetaria es: {media_poblacional:.2f}")

if __name__ == "__main__":
    os.makedirs("graficas", exist_ok=True)
    fun_planetas()